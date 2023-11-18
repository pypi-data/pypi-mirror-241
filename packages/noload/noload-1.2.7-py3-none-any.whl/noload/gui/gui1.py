import tkinter as tk
from tkinter import ttk,TOP,BOTH
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
import sys,pandas
import numpy as np
from openpyxl.styles import Font, Color
from openpyxl import Workbook
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg,\
    NavigationToolbar2Tk
### Tk App ###

class App:
    def __init__(self, root, options, df, specifications):
        self.root = root
        self.options = options
        self.selected_options = []  # Keep track of selected options
        self.df = df
        self.specifications = specifications

        self.create_widgets()

    def create_widgets(self):

        # Create a Treeview widget
        self.treeview = ttk.Treeview(self.root, columns=("Variable", "Min",
                    "Max", "Value", "Type", "Input/Output"), show="headings",
                            height=len(self.specifications))
        self.treeview.pack(padx=10, pady=5)

        # Add column headings
        self.treeview.heading("Variable", text="Variable")
        self.treeview.heading("Min", text="Min")
        self.treeview.heading("Max", text="Max")
        self.treeview.heading("Value", text="Value")
        self.treeview.heading("Type", text="Type")
        self.treeview.heading("Input/Output", text="Input/Output")

        # Define color tags based on type and min/max columns
        type_colors = {
            #"bounds": "black",
            "ineq_cstr": "grey",
            "objective": "red",
            "eq_cstr": "green",
            #"free": "black",
        }

        type_names = {
            "bounds": "constrained",
            "ineq_cstr": "inequality",
            "objective": "objective",
            "eq_cstr": "fixed",
            "free": "free",
        }

        # Add options and additional information
        for option in self.options:
            type_ = self.specifications.Type[option]
            value = self.specifications.Value[option]
            input_output = self.specifications.In_Out[option]

            if type_=='bounds' or type_=='ineq_cstr':
                if isinstance(value[0],list):
                    min_value, max_value=str(np.transpose(np.array(value))[0]),\
                                          str(np.transpose(np.array(value))[1])
                else:
                    min_value, max_value = value
                value_text = ""

            elif type_=='eq_cstr':
                min_value = ""
                max_value = ""
                value_text = str(value)

            else:
                min_value = ""
                max_value = ""
                value_text = ""


            # Add rows and colors
            self.treeview.insert("", "end", values=(option, min_value,
               max_value, value_text, type_names.get(type_), input_output),
                                 tags=(type_,))
            self.treeview.tag_configure(type_, foreground=
            type_colors.get(type_, "black"))

        # Create a scrollbar for the Treeview
        scrollbar = ttk.Scrollbar(self.root, orient="vertical",
                                  command=self.treeview.yview)
        self.treeview.configure(yscroll=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

        # Create a context menu for the Treeview
        self.option_menu = tk.Menu(self.root, tearoff=False)
        self.option_menu.add_command(label="Additional Info",
                                     command=self.show_additional_info)
        self.option_menu.entryconfigure(0, state="disabled")  # Disable the "Additional Info" menu item initially
        self.treeview.bind("<Button-3>", self.show_context_menu)

        self.export_button = tk.Button(self.root, text="Export",
                                       command=self.export_to_excel)
        self.export_button.pack(pady=5)
        self.canvas=None
        self.toolbar=None
        self.exit_button = tk.Button(self.root, text="Exit",
                                     command=self.exit_program)
        self.exit_button.pack(pady=5)

    def show_context_menu(self, event):
        selection = self.treeview.selection()
        if selection:
            self.selected_options = [self.options[self.treeview.index(item)]
                                     for item in selection]  # Update selected options
            menu_label = f"Additional Info ({len(self.selected_options)} " \
                         f"options selected)"
            self.option_menu.entryconfigure(0, label=menu_label)
            self.option_menu.entryconfigure(0, state="normal")  # Enable the "Additional Info" menu item
            if len(self.selected_options) > 1:
                self.option_menu.entryconfigure(1, label="Plot",
                                    command=self.plot_multiple_selected_options)
            else:
                self.option_menu.entryconfigure(1, label="Plot",
                                    command=self.plot_one_selected_option)
            self.option_menu.tk_popup(event.x_root, event.y_root)

    def show_additional_info(self):
        for selected_option in self.selected_options:
            print(f"Additional info for {selected_option}")


    def plot_multiple_selected_options(self):
        self.fig, self.ax = plt.subplots()
        for selected_option in self.selected_options:
            self.ax.plot(self.df['IterationNumber'], self.df[selected_option])
            self.ax.scatter(self.df['IterationNumber'],self.df[selected_option],
                       label=f'{selected_option}')

            if len(self.selected_options) <= 3:
            # Label each dot
                for iteration, x, y in zip(self.df['IterationNumber'],
                        self.df['IterationNumber'], self.df[selected_option]):
                    self.ax.annotate(iteration,(x,y),textcoords="offset points",
                                xytext=(0, 10), ha='center', va='bottom')

            self.ax.legend()
        self.ax.grid()
        self.ax.set_xlabel('Iteration Number')
        self.ax.set_ylabel('Option Values')
        self.ax.xaxis.set_major_locator(MaxNLocator(integer=True))
        self.ax.ticklabel_format(useOffset=False, style='plain')

        if self.canvas!=None:
            self.canvas.get_tk_widget().pack_forget()
            self.root.winfo_children()[-1].pack_forget()
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.root)  # A tk.DrawingArea.
        self.canvas.get_tk_widget().pack(side=TOP, fill=BOTH, expand=1)
        self.toolbar=NavigationToolbar2Tk(self.canvas, self.root)
        self.canvas.draw_idle()

    def export_to_excel(self):
        # Create a Workbook object
        workbook = Workbook()

        # Create a new sheet
        sheet = workbook.active

        # Get the columns to export excluding 'IsBestSolution', 'IsSolution', and 'pareto_pts'
        export_columns = [col for col in self.df.columns if col not in
                          ['IsBestSolution', 'IsSolution', 'pareto_pts']]

        # Write the column names to the sheet
        sheet.append(export_columns)

        # Write the row values to the sheet
        for _, row in self.df[export_columns].iterrows():
            values = row.tolist()
            sheet.append(values)

        # Apply conditional text color to the desired columns
        for column in self.specifications.index:
            current_spec = self.specifications.Type[column]
            if current_spec == "bounds":
                spec_values = self.specifications.Value[column]
                min_value, max_value = spec_values[0], spec_values[1]
                font_color = "000000"  # Default black color
                for index, value in self.df[column].items():
                    if value <= min_value:
                        font_color = "0000FF"  # Blue color
                    elif value >= max_value:
                        font_color = "FF0000"  # Red color
                    if column in export_columns:
                        cell = sheet.cell(row=index+2,
                                          column=export_columns.index(column)+1)
                        cell.font = Font(color=Color(rgb=font_color))
                        cell.value = value
            elif current_spec == "eq_cstr":
                font_color = "808080"  # Grey color
                for index in self.df.index:
                    if column in export_columns:
                        cell = sheet.cell(row=index+2,
                                          column=export_columns.index(column)+1)
                        cell.font = Font(color=Color(rgb=font_color))
                        cell.value = self.df[column][index]
            elif current_spec == "objective":
                font_color = "00FF00"  # Green color
                for index in self.df.index:
                    if column in export_columns:
                        cell = sheet.cell(row=index+2,
                                          column=export_columns.index(column)+1)
                        cell.font = Font(color=Color(rgb=font_color))
                        cell.value = self.df[column][index]

        # Save the workbook to a file
        workbook.save("exported_data.xlsx")
        print("Result saved!")

    def plot_one_selected_option(self):

        selected_option = self.selected_options[0]
        self.fig, self.ax = plt.subplots()

        if isinstance(self.df[selected_option][0],list):
            index=[selected_option+str(i) for i in range(len(self.df
                                                [selected_option][0]))]
            df=pandas.DataFrame(np.array(self.df[selected_option].values.
                                         tolist()),columns=index)

            for element in index:
                self.ax.plot(self.df['IterationNumber'],df[element],
                             label=element)
                self.ax.scatter(self.df['IterationNumber'], df[element])

            # Label each dot
                for iteration, x, y in zip(self.df['IterationNumber'],
                    self.df['IterationNumber'],df[element]):
                    self.ax.annotate(iteration,(x,y),textcoords="offset points",
                            xytext=(0, 10), ha='center', va='bottom')

            current_spec = self.specifications.Value[selected_option]
            current_type = self.specifications.Type[selected_option]

            for element in index:
                current_spec_el=current_spec[index.index(element)]
                if current_type=="bounds":
                    self.ax.axhline(y=current_spec_el[1], color='red',
                                linestyle='--',label=element+'_max')
                    self.ax.axhline(y=current_spec_el[0],color='blue',
                                linestyle='--',label=element+'_min')

                elif current_type=='ineq_cstr':
                    if current_spec_el[1]!=None:
                        self.ax.axhline(y=current_spec_el[1],color='black',
                            linestyle='--',label=element + '_max')
                    if current_spec_el[0]!=None:
                        self.ax.axhline(y=current_spec_el[0], color='grey',
                            linestyle='--',label=element + '_min')

                elif current_type=='eq_cstr':
                    self.ax.axhline(y=current_spec_el, color='black',
                            linestyle='--',label=element+'_cstr')

        else:
            self.ax.plot(self.df['IterationNumber'],self.df[selected_option],
                    label=selected_option)
            self.ax.scatter(self.df['IterationNumber'],self.df[selected_option])

            # Label each dot
            for iteration, x, y in zip(self.df['IterationNumber'],
                        self.df['IterationNumber'], self.df[selected_option]):
                self.ax.annotate(iteration, (x, y), textcoords="offset points",
                            xytext=(0, 10), ha='center', va='bottom')

            current_spec = self.specifications.Value[selected_option]
            current_type = self.specifications.Type[selected_option]

            if current_type=="bounds":
                self.ax.axhline(y=current_spec[1], color='red', linestyle='--',
                           label=selected_option+'_max')
                self.ax.axhline(y=current_spec[0], color='blue', linestyle='--',
                           label=selected_option+'_min')

            elif current_type=='ineq_cstr':
                if current_spec[1]!=None:
                    self.ax.axhline(y=current_spec[1], color='black',
                        linestyle='--',label=selected_option + '_max')
                if current_spec[0]!=None:
                    self.ax.axhline(y=current_spec[0], color='grey',
                        linestyle='--', label=selected_option + '_min')

            elif current_type=='eq_cstr':
                self.ax.axhline(y=current_spec, color='black', linestyle='--',
                           label=selected_option+'_cstr')

        self.ax.grid()
        self.ax.set_xlabel("IterationNumber")
        self.ax.set_title(selected_option)

        # Add the legend
        self.ax.legend()

        self.ax.xaxis.set_major_locator(MaxNLocator(integer=True))
        self.ax.ticklabel_format(useOffset=False, style='plain')

        if self.canvas!=None:
            self.canvas.get_tk_widget().pack_forget()
            self.root.winfo_children()[-1].pack_forget()
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.root)  # A tk.DrawingArea.
        self.canvas.get_tk_widget().pack(side=TOP, fill=BOTH, expand=1)
        self.toolbar = NavigationToolbar2Tk(self.canvas, self.root)
        self.canvas.draw_idle()

    def exit_program(self):
        sys.exit()
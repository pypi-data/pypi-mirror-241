import time
from typing import Sequence, Type
import os
import functools

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.cell.cell import Cell
from win32com.client import DispatchEx


def just_open(filename):
    '''
    openpyxl 保存的 excel 文件如果没有打开保存过，公式的值不会被计算
    '''

    xlApp = DispatchEx("Excel.Application")

    xlApp.Visible = False

    xlBook = xlApp.Workbooks.Open(filename)

    xlBook.Save()

    xlBook.Close()


class Column(str):
    '''
    代表 excel 表中的列名
    '''
    attr_name: str
    choices: Sequence

    def __new__(cls, val='', choices=None):
        obj = str.__new__(cls, val)
        obj.choices = choices
        return obj

    def __set_name__(self, cls, name):
        self.attr_name = name

    def __get__(self, obj, cls):
        if obj is None:
            return self

        index = obj.__dict__.get(self.attr_name)
        if index is None:
            raise AttributeError

        if obj.row is None:
            return index
        return obj.row[index]

    def __set__(self, obj, value):
        if obj.row is None:
            raise AttributeError('unbind object can not assign')

        index = obj.__dict__.get(self.attr_name)
        if index is None:
            raise AttributeError

        obj.row[index] = value

    def __delete__(self, instance):
        pass


class Index(int):
    column: Column

    def __new__(cls, val, col=None):
        obj = int.__new__(cls, val)
        obj.column = col
        return obj


class TableType(type):
    def __new__(cls, name, bases, attrs: dict):
        if not bases:
            return type.__new__(cls, name, bases, attrs)

        # 将 `__annotations__` 中的字段转换成 `class` 属性
        # 每个属性的值是 `Column` 类型的实例
        for attr, _ in attrs['__annotations__'].items():
            if attr.startswith('__'):
                continue

            col: Column = attrs.pop(attr, None)
            if col is None:
                attrs[attr] = Column(val=attr)
            else:
                if not col:
                    new_col = Column(val=attr)
                    new_col.__dict__.update(col.__dict__)
                    attrs[attr] = new_col
                else:
                    attrs[attr] = col

        return type.__new__(cls, name, bases, attrs)


class Table(metaclass=TableType):
    __sheetname__: str = ''

    def __init__(
            self, sheet: Worksheet, header_row=1, header_all_exist=False
    ) -> None:
        # 从 `sheet` 中取第一行作为表头
        heads = [c.value for c in sheet[header_row]]

        if header_all_exist:
            def header_not_found(e):
                raise e
        else:
            def header_not_found(e):
                pass

        for attr, col in self.__class__.__dict__.items():
            if not isinstance(col, Column):
                continue

            try:
                self.__dict__[attr] = Index(heads.index(col), col=col)
            except ValueError as e:
                header_not_found(e)
        self.sheet = sheet
        self.row = None

    @property
    def columns(self) -> Sequence[Column]:
        '''
        获取当前sheet有的列，从实例对象中获取
        '''

        cols = []

        for i in self.__dict__.values():
            if not isinstance(i, Index):
                continue

            cols.append(
                i.column
            )
        return cols

    @property
    def columns_count(self) -> int:
        return len(self.indexes)

    def new_orphan_row(self, empty=None) -> list:
        '''
        根据当前列数量创建一行，该行还没有添加到表中
        '''
        return [empty] * self.columns_count

    def new_orphan_row_and_bind(self, empty=None) -> list:
        row = self.new_orphan_row(empty)
        self.bind_row(row)
        return row

    def append_bind_orphan_row(self):
        self.sheet.append(self.row)

    @classmethod
    def get_full_columns(cls) -> Sequence[str]:
        cols = []
        for _, val in cls.__dict__.items():
            if not isinstance(val, Column):
                continue
            cols.append(val)
        return cols

    @property
    def indexes(self) -> Sequence[int]:
        rst = []
        for v in self.__dict__.values():
            if not isinstance(v, Index):
                continue
            rst.append(v)
        return rst

    @classmethod
    def write_full_heads(cls, sheet: Worksheet):
        '''
        将该 model 上定义的字段全部写入到 sheet
        '''
        sheet.append(cls.get_full_columns())

    def write_heads2other_sheet(self, sheet: Worksheet):
        '''
        将该 Table 绑定的字段写到另外一个 sheet 中

        字段可能少于 write_full_fields，只包含实例化该对象时，传入的 sheet 里有的字段
        '''
        sheet.append(self.columns)

    def append_row(self, src_row: Sequence):
        '''
        给当前绑定的 sheet 添加一行数据
        '''
        self.sheet.append(src_row)

    def append_bindrow2other_sheet(self, dst_sheet: Worksheet):
        if not self.row:
            raise ValueError('no bind row')
        self.append_row2other_sheet(dst_sheet, self.row)

    def append_row2other_sheet(self, dst_sheet: Worksheet, src_row: Sequence):
        '''
        将当前 Table 绑定的 sheet 中的行，添加到另外的表中。
        相当于将当前绑定表的 Table 上的字段添加到另外的表中
        '''
        new_row = []

        for i in self.indexes:
            new_row.append(src_row[i])
        dst_sheet.append(new_row)

    def copy_bind_row(self, other: 'Table'):
        for col in other.columns:
            setattr(other, col.attr_name, getattr(self, col.attr_name, None))

    def bind_row(self, row):
        self.row = row

    def unbind_row(self):
        self.row = None


class Excel:

    def __init__(self, file: str, data_only=False) -> None:
        self.file: str = file
        self.workbook: Workbook = None
        self.worksheet: Worksheet = None
        self.data_only = data_only

        self.load_workbook()

    def close(self):
        if self.workbook:
            self.workbook.close()

    def save(self, compute_formula=True, save_as=None):
        save_as = save_as or self.file
        self.workbook.save(save_as)
        abspath = os.path.abspath(self.file)

        if compute_formula:
            just_open(abspath)

    def load_workbook(self) -> Workbook:
        print(f'loading <{self.file}> ...')
        t1 = time.time()
        wb = load_workbook(self.file, data_only=self.data_only)
        t2 = time.time()
        print(f'load <{self.file}> ok, cost: {t2 - t1}s')
        self.workbook = wb
        return wb

    def new_workbook(self) -> Workbook:
        self.workbook = Workbook()
        return self.workbook

    def new_work_book_with_a_sheet(self, title: str = None) -> Worksheet:
        self.worksheet = self.new_workbook().active
        if title:
            self.worksheet.title = title
        return self.worksheet

    def select_worksheet(self, title: str) -> Worksheet:
        self.worksheet = self.workbook[title]
        return self.worksheet

    def select_active_worksheet(self) -> Worksheet:
        return self.workbook.active

    def create_sheet_from_table_cls(
            self, table_cls: Type[Table], delete_old=False, write_header=True
    ) -> Worksheet:
        sheet = self.create_sheet(
            table_cls.__sheetname__, delete_old=delete_old
        )
        if write_header:
            table_cls.write_full_heads(sheet)
        return sheet

    def create_sheet(self, title: str, delete_old=False) -> Worksheet:
        if self.workbook is None:
            self.load_workbook()

        if delete_old:
            try:
                ws = self.workbook[title]
                self.workbook.remove(ws)
            except KeyError:
                pass

        ws = self.workbook.create_sheet(title)
        return ws

    @classmethod
    def insert_columns(
        cls, sheet: Worksheet, pos: int, cols: Sequence[str], not_exist=True
    ):
        if not_exist:
            heads = [c.value for c in sheet[1]]
            for col in cols:
                if col in heads:
                    return

        for col in cols:
            sheet[coordinate(1, pos)] = col
            pos += 1

    @classmethod
    def append_columns(
        cls, sheet: Worksheet, cols: Sequence[str], not_exist=True
    ):
        insert_col = sheet.max_column + 1
        cls.insert_columns(sheet, insert_col, cols, not_exist=not_exist)


def num2excel_col(num) -> str:
    '''
    将数值形式的列转换成 excel 字母形式的列

    >>> num2excel_col(30)
    'AD'
    '''

    rst = ""
    while num > 0:
        num -= 1
        rst = chr((num % 26) + ord('A')) + rst
        num //= 26
    return rst


def coordinate(row, col) -> str:
    return f'{num2excel_col(col)}{row}'


def formula(*args):
    items = ['=']

    for arg in args:
        if isinstance(arg, str):
            items.append(arg)

        if isinstance(arg, Cell):
            items.append(arg.coordinate)
    return ' '.join(items)


def nonnegative_number(raw_formula: str):
    raw_formula = raw_formula[1:]
    formula = f'=if({raw_formula}>0, {raw_formula}, 0)'
    return formula


def if_error(raw_formula: str, errdefault=None):
    raw_formula = raw_formula[1:]
    formula = f'=if(iserror({raw_formula}), {errdefault}, {raw_formula})'
    return formula


def if_error_factory(errdefault):
    return functools.partial(if_error, errdefault=errdefault)


def formula_format(
    format_spec: str, *args, iserror=True, errdefault=0
) -> str:
    '''
    iserror=True, errdefault=0 这两个参数准备废弃，使用 if_error 函数代替
    只是还有大量老代码在使用
    '''

    new_args = []

    for arg in args:
        if isinstance(arg, Cell):
            new_args.append(arg.coordinate)
            continue

        new_args.append(arg)

    f = format_spec.format(*new_args)
    if iserror:
        f = f'if(iserror({f}), {errdefault}, {f})'
    return f'={f}'


def set_formula_format(
    cell: Cell, format_spec: str, *args, wrappers: Sequence = None
):
    f = formula_format(
        format_spec, *args, iserror=False
    )

    if wrappers:
        for w in wrappers:
            f = w(f)
    cell.value = f

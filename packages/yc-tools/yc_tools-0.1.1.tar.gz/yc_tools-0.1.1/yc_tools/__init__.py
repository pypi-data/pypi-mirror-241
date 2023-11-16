from asyncore import write
from multiprocessing import Lock
import os
import time
from time import sleep
import openpyxl
import winsound
import logging
import logging.handlers
from logging import *
import threading
import datetime
import serial
from queue import Queue
import threading
class yc_serial():
    def __init__(self,port,baudrate):
        self.lock = threading.RLock()
        self.log = LOG("serial")
        self.q_receive = Queue()
        self.port1 = serial.Serial(port,baudrate)
        if (self.port1.is_open):
            self.log.info("打开端口:%s成功"%port)
        else:
            self.log.info("打开端口:%s失败"%port)
        # Creat_Thread(self.receivemsg)

    def write(self,data):
        self.lock.acquire()
        self.port1.write(data)
        self.lock.release()
    def receivemsg(self):
        while True:
            size=self.port1.in_waiting
            if size:
                self.recdata = self.port1.readline().decode('utf-8').replace("\r","").replace("\n","")
                if self.recdata != "":
                    self.log.info(str(self.recdata))
                    self.q_receive.put(self.recdata)

    def write_str(self,senddata:str):
        senddata1 = senddata.encode('utf-8')
        self.write(senddata1)
        self.log.info("write->"+senddata)

    def write_hex(self,senddata:list):
        self.write(senddata)
        self.log.info("write->"+str(senddata))

class LOG(logging.Logger):
    def __init__(self,
                 name='root',
                 level='DEBUG'
                 ):
        now_date = datetime.datetime.now().strftime("%Y-%m-%d")
        file="./logs/"+now_date+"_"+name+".log"
        filepath = os.path.split(file)[0]
        if not os.path.exists(filepath):
            os.makedirs(filepath)
        super().__init__(name)
        self.setLevel(level)
        fmt = logging.Formatter("%(asctime)s %(name)s-line:%(lineno)d-%(levelname)s:     %(message)s", "%H:%M:%S")
        file_handler = logging.FileHandler(file,encoding='utf-8')
        file_handler.setLevel(level)
        file_handler.setFormatter(fmt)
        self.addHandler(file_handler)
        stream_handler = logging.StreamHandler()
        stream_handler.setLevel(level)
        stream_handler.setFormatter(fmt)
        self.addHandler(stream_handler)
    def info(self, msg, *args, **kwargs):
        """
        Log 'msg % args' with severity 'INFO'.

        To pass exception information, use the keyword argument exc_info with
        a true value, e.g.

        logger.info("Houston, we have a %s", "interesting problem", exc_info=1)
        """
        if self.isEnabledFor(INFO):
            # beep_on()
            self._log(INFO, msg, args, **kwargs)

def mkdir(path):
    folder = os.path.exists(path)
    if not folder:                   #判断是否存在文件夹如果不存在则创建为文件夹
        os.makedirs(path)            #makedirs 创建文件时如果路径不存在会创建这个路径

def Creat_Thread(thread,*args1):
    thread_uart_read=threading.Thread(target=thread,args=args1)
    thread_uart_read.setDaemon(True)
    thread_uart_read.start()

def get_cmd_result(cmd):
    # print(cmd)
    result = os.popen(cmd)
    res = result.read()
    for line in res.splitlines():
        print (line)
    time.sleep(0.1)
    return res
def xw_toExcel(data,excelname):  # xlsxwriter库储存数据到excel
    try:
        workbook = openpyxl.load_workbook(excelname)  # 打开工作簿
        worksheet1 = workbook.active  # 创建子表
    except Exception as a:
        print(a)
        workbook = openpyxl.Workbook(excelname)
        workbook.save(filename= excelname)  # 关闭表
        workbook.close()
        workbook = openpyxl.load_workbook(excelname)  # 打开工作簿
        worksheet1 = workbook.active
    rows = worksheet1.max_row
    j = 0
    for i in data:
        j+=1
        worksheet1.cell(row=rows+1,column = j).value = i
    workbook.save(filename= excelname)  # 关闭表
    workbook.close()
def xw_columnToExcel(data,excelname):  # xlsxwriter库储存数据到excel
    try:
        workbook = openpyxl.load_workbook(excelname)  # 打开工作簿
        worksheet1 = workbook.active  # 创建子表
    except Exception as a:
        print(a)
        workbook = openpyxl.Workbook(excelname)
        workbook.save(filename= excelname)  # 关闭表
        workbook.close()
        workbook = openpyxl.load_workbook(excelname)  # 打开工作簿
        worksheet1 = workbook.active
    columns = worksheet1.max_column
    j = 0
    for i in data:
        j+=1
        worksheet1.cell(row=j,column = columns+1).value = i
    workbook.save(filename= excelname)  # 关闭表
    workbook.close()
def e_pu():
    while('CPU Stopped' not in get_cmd_result("e pu")):
        time.sleep(0.05)
        continue
    return
def e_p():
    while('CPU Stopped' not in get_cmd_result("e p")):
        time.sleep(0.05)
        continue
    return
def e_cu():
    while('CPU Running' not in get_cmd_result("e cu")):
        continue
    return
def e_c():
    while('CPU Running' not in get_cmd_result("e c")):
        continue
    return
def e_ku():
    while('CPU B reseted' not in get_cmd_result("e k")):
        continue
    return
def e_k():
    while('cm0 reseted' not in get_cmd_result("e k")):
        continue
    return
def count_bits_2(value):
    count = 0
    while(value):
        value &= value - 1
        count += 1
    return count
def to_hex(val):
    return hex(val).replace("0x","")

def beep_on():
    duration = 1000  # millisecond
    freq = 440  # Hz
    winsound.Beep(freq, duration)
def set_environ(var,val):
    os.environ[var]=val

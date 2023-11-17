import pandas as pd
import openpyxl
from openpyxl.styles import NamedStyle, Font, Alignment, numbers
from openpyxl.styles import Alignment, PatternFill

from ExcelPar.mod.SetGlobal import SetGlobal

class CreateLeadReport:
    @classmethod
    def CreateLeadReport(cls, tb:pd.DataFrame, tb_월별:pd.DataFrame) -> str: #return LeadFileName

        #### 총괄 분석적검토 자동화 파일을 생성합니다.
        #tb_월별 : 전역변수 Read
        분석적검토 = tb_월별[['T1', 'T2', 'T3', 'T4','Company code','통제활동의존', '위험수준','계정과목코드', '계정과목명', 'CY', 'PY', 'PY1','증감금액', '증감비율', 'Threshold', '분석대상']]
        분석적검토['PY'] = tb['PY'] #누적합을 위해 구현한 분반기코드 원복 => PY는 TB PY로
        분석적검토1 = 분석적검토[분석적검토["분석대상"] == "O"]
        # 분석적검토

        분석적검토1.loc[:,"주요증감"] = ""
        분석적검토1.loc[:,"Refer to 주요증감_파일"] = ""

        분석적검토1.loc[:,"주요증감"]= SetGlobal.검토문장 #Global
        #231025 DEBUG : 만약, TB COA에 중복이 있는 경우 여기서 걸림. 따라서 TB COA에 중복이 있으면 안됨
        
        분석적검토1.loc[:,"Refer to 주요증감_파일"] = "분석보고서_" + SetGlobal.분석계정과목 + ".xlsx"

        # 분석적검토와 분석적검토1을 합치기
        분석적검토 = pd.concat([분석적검토, 분석적검토1], ignore_index=True)

        # 분석적검토1의 "주요증감" 열을 "검토문장" 값으로 채우기
        분석적검토["주요증감"].fillna(분석적검토1["주요증감"], inplace=True)
        분석적검토["Refer to 주요증감_파일"].fillna(분석적검토1["Refer to 주요증감_파일"], inplace=True)

        분석적검토 = 분석적검토.drop_duplicates()

        print("총괄파일을 생성합니다.")
        #global LeadFileName
        LeadFileName = f"총괄_분석적검토_자동화_{SetGlobal.ClientNameDate}_{SetGlobal.Level}.xlsx"
        분석적검토.to_excel(LeadFileName)

        # 엑셀 파일 열기
        workbook = openpyxl.load_workbook(LeadFileName)
        worksheet = workbook.active  # 또는 원하는 시트를 선택합니다.

        # pandas 데이터프레임의 열마다 컴마 스타일 적용
        for col_num, column in enumerate(분석적검토.columns, start=1):
            for cell in worksheet.iter_rows(min_row=2, min_col=col_num, max_col=col_num):
                for cell in cell:
                    cell.number_format = '#,##0'
                    
        for row in range(1, 1500):
                worksheet["O{}".format(row)].number_format = '#,##0.00'
                

        for column_letter in range(ord('B'), ord('P') + 1):
            column_letter = chr(column_letter)
            worksheet.column_dimensions[column_letter].width = 15

        worksheet.column_dimensions['R'].width = 50

        # 행 삽입하기
        worksheet.insert_rows(1,4)

        # formatting the report
        worksheet['A1'] = 'Preliminary Analytical Procedures'
        worksheet['A2'] = "(유의사항: 분석적 검토를 지원하기 위해 자동화된 방법으로 원장 내용을 추출하여 수행한 결과이므로 참고용으로 활용하시기 바랍니다.)"
        worksheet['A1'].font = Font('Arial', bold=True, size=18)
        worksheet['A2'].font = Font('Arial', bold=True, size=12)

        # 회사명 / M / PM / CTT
        worksheet.title ='총괄'

        worksheet['I3'] = '회사명'
        worksheet['I4'] = SetGlobal.ClientNameDate
        worksheet['J3'] = '계정과목수준'
        worksheet['J4'] = SetGlobal.Level
        worksheet['K3'] = 'PM'
        worksheet['K4'] = SetGlobal.PM
        worksheet['L3'] = 'CTT'
        worksheet['L4'] = SetGlobal.De_minimis
        
        worksheet['Q2'] = '분반기검토시 PY금액은 BS의 경우 전기말, PL의 경우 전년동기말입니다. (PY1은 일괄 전전기말)'
        worksheet['Q3'] = 'Threshold = Min[0.2CY, 0.5PM]' #231025 typo
        worksheet['Q4'] = '분석대상 = 증감금액 >= Threshold | 증감비율 >= 20%'

        for row in worksheet['I3:L4']:
            for cell in row:
                cell.fill = PatternFill(start_color='666699', end_color='666699',fill_type='solid')
                cell.alignment = Alignment(horizontal="center")

        for row in worksheet['I3:L3']:
            for cell in row:
                cell.fill = PatternFill(start_color='666699', end_color='666699',fill_type='solid')
                cell.alignment = Alignment(horizontal="center")

        for row in worksheet['K4:L4']:
            for cell in row:
                cell.number_format = '#,###'    

        # 조건부 서식을 적용할 열 범위 지정 (예: F5부터 F200까지)
        cell_range = worksheet['Q5:Q1000']

        # 빨간색 배경에 하얀색 폰트 스타일 설정
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # 빨간색 배경색
        white_font = Font(color="FFFFFF")  # 하얀색 폰트

        # 조건부 서식 설정
        for row in cell_range:
            for cell in row:
                if cell.value == "O":
                    cell.fill = red_fill  # 빨간색 배경색 적용
                    cell.font = white_font  # 하얀색 폰트 적용
                    
        worksheet.column_dimensions['S'].width = 40

        worksheet.freeze_panes = 'B6'  # B6 셀을 기준으로 위쪽 행과 왼쪽 열을 틀고정
        worksheet.column_dimensions.group('B','H',hidden=True)     # B 부터 D 열 숨기기


        # B5부터 S5까지의 범위 선택
        cell_range = worksheet['B5:S5']

        # 검은색 배경에 흰색 폰트색 설정
        black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # 검은색 배경색
        white_font = Font(color="FFFFFF", bold=True)  # 흰색 볼드 폰트

        # 범위에 스타일 적용
        for row in cell_range:
            for cell in row:
                cell.fill = black_fill  # 검은색 배경색 적용
                cell.font = white_font  # 흰색 폰트 적용
                
        workbook.save(LeadFileName)
        workbook.close()

        return LeadFileName

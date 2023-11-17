#전역부
import ExcelPar.mylib.myFileDialog as myfd
import os
import glob
import pandas as pd
import time
import gc
import openpyxl

###########################################
#경로 내 엑셀파일을 합치는 코드
#주. 너무 느리면 엑셀 자체를 text로 추출한 뒤에 사용하는게 빠름

#####################################
#통합본
def CircleSheets(filename:str, sheetCount:int) -> pd.DataFrame:
    #filename = myfd.askopenfilename()

    #파일을 읽어서 시트 수를 찾아낸다
    #wb = openpyxl.load_workbook(filename)
    #sheetCount = len(wb.sheetnames)

    dfCon = pd.DataFrame()

    for i in range(sheetCount):
        tmpTime = time.time()
        print(i,"번째 시트 순환시작")
        df = pd.read_excel(filename,sheet_name=i, skiprows=[0]) #Should be i
        print("sheet Load 소요시간: ",time.time()-tmpTime)
        dfCon = pd.concat([dfCon, df])        
        print(i,"번째 시트 순환완료")
    return dfCon


def CircleFiles():
    path = myfd.askdirectory() #동적으로 폴더 설정
    list = glob.glob(path + "/" + "*.xls*")

    print(len(list))

    #합산
    dfCon = pd.DataFrame()

    startTime = time.time()
    for i in list:    
        nowTime = time.time()
        print(i," 시작:")

        #시트수 읽기
        wb = openpyxl.load_workbook(i)
        sheetCount = len(wb.sheetnames)

        if sheetCount == 1:
            #df = pd.read_excel(i)
            df = pd.read_excel(i,skiprows=[0]) #ignore first row
            df.columns = [i.strip() for i in df.columns] #추가코드. strip (TRIM)
        else: #복수 시트인 경우
            df = CircleSheets(i, sheetCount)

        dfCon = pd.concat([dfCon,df])
        print(i," 종료:")
        print("소요시간 : ",time.time() - nowTime)

    print("총 소요시간: ",time.time()-startTime)
    return dfCon
    #dfCon.to_pickle("JEJU2023.pickle")

def RunConcatExcelFiles():    
    df = CircleFiles()
    fileName = input("저장할 파일명(Pickle)>>")
    df.to_pickle(fileName)
    print("DONE")

if __name__=='__main__':
    RunConcatExcelFiles()

#####################################



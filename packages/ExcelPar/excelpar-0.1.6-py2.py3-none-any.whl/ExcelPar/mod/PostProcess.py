import os
from openpyxl import load_workbook
from ExcelPar.mylib.create_and_move_folder import create_and_move_folder

from ExcelPar.mod.SetGlobal import SetGlobal

class Postprocess:
    @classmethod
    def Postprocess(cls, LeadFileName : str): #Level은 Detail이 기본값임

        # Excel 파일을 '분석보고서_Reporting' 폴더로 이동 => 마지막에 수행
        #create_and_move_folder('.xlsx', '분석보고서_Reporting', '분석보고서_분석적검토_자동화')

        # 월별 변동액/잔액 파일을 '분석보고서_Monthly' 폴더로 이동
        create_and_move_folder('.xlsx', '분석보고서_Monthly', '분석보고서_계정별월별')

        # Excel 파일을 '분석보고서_Reporting_세부' 폴더로 이동
        create_and_move_folder('.xlsx', '분석보고서_Reporting_세부', '분석보고서')

        # Excel 파일 로드
        FileName = LeadFileName
        wb = load_workbook(FileName)
        ws = wb.active  # 첫 번째 워크시트

        # 분석보고서_Reporting_세부 폴더의 파일 리스트 가져오기 ## 세부 폴더를 추가해서 올려야함 >>>>> PHW    
        reporting_folder_path = "분석보고서_Reporting_세부"
        file_names = os.listdir(reporting_folder_path)

        # S열 6행부터 200행까지 검사
        for row in range(5, 1500):
            cell_value = ws[f"S{row}"].value  # S열의 값을 가져옴
            if cell_value in file_names:  # 값이 폴더의 파일 이름과 동일한 경우
                ws[f"S{row}"].hyperlink = os.path.join(reporting_folder_path, cell_value)  # 링크 추가

        wb.save(FileName) 
        wb.close()

        # Excel 파일을 '분석보고서_Reporting' 폴더로 이동 => 마지막에 수행
        create_and_move_folder('.xlsx', '분석보고서_Lead', '총괄_분석적검토_자동화')

        Folder1 = "분석보고서_Monthly"
        Folder2 = "분석보고서_Reporting_세부"
        Folder3 = "분석보고서_Lead"
        
        try:
            os.rename(Folder1,Folder1 + "_" + SetGlobal.Level)    
        except:
            print("폴더명 변경 오류. 이미 폴더가 있는 것 같습니다.")
        try:
            os.rename(Folder2,Folder2 + "_" + SetGlobal.Level)    
        except:
            print("폴더명 변경 오류. 이미 폴더가 있는 것 같습니다.")
        try:
            os.rename(Folder3,Folder3 + "_" + SetGlobal.Level)
        except:
            print("폴더명 변경 오류. 이미 폴더가 있는 것 같습니다.")

    #####################################################################
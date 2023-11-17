import pandas as pd
import numpy as np
import openpyxl
import dask.dataframe as dd

from ExcelPar.mod.SetGlobal import SetGlobal
from ExcelPar.mylib.TimeCheck import TimeCheck as tc

class SummarizeMonthlyVarAmount:
    @classmethod
    def SummarizeMonthlyVarAmount(cls, gl:dd.DataFrame, tb:pd.DataFrame) -> pd.DataFrame: #Return : tb_월별

        #############################################################

        ### 1. 월별 순변동액 (not 잔액)

        #a. 당기GL을 추출한다. => 여기 고쳐야 함 dask로
        gl_당기:dd.DataFrame = gl[gl["연도"] == "CY"] 

        #b. 당기GL을 행은 계정과목 / 열은 회계월로 피벗(SUM)한다. => TB와 JOIN => 여기 고쳐야함 dask로
        #gl_월별 = pd.pivot_table(gl_당기,values=('전표금액'),index=['계정과목코드','계정과목명'],columns=['회계월'],aggfunc=np.sum)
        tc.Set()
        gl_월별Tmp:pd.DataFrame|pd.DataFrame = gl_당기.groupby(['계정과목코드','계정과목명','회계월'])['전표금액'].sum()#.compute #UNSTACK IS ONLY PANDAS
        if SetGlobal.bDask: gl_월별Tmp = gl_월별Tmp.compute()
        tc.Check("gl_당기.groupby")
        
        gl_월별:pd.DataFrame = gl_월별Tmp.unstack('회계월')

        # 후처리
        gl_월별 = gl_월별.fillna(0)

        #a. 월별증감액    
        Filename = f'분석보고서_계정별월별증감액_{SetGlobal.ClientNameDate}_{SetGlobal.Level}.xlsx'
        gl_월별.to_excel(Filename)

        # 안내문 추가        
        wb = openpyxl.load_workbook(Filename)
        ws = wb.active
        ws.insert_rows(1)
        ws['A1'] = "#계정과목별 월별 순변동금액 - G/L로부터 추출한 금액"    
        wb.save(Filename)
        wb.close()

        print("월별 순변동액 보고서를 생성하였습니다.")

        #############################################################

        ### 2. 월말 잔액 (누적변동액)
        #a. GL은 위에서 추출한 gl_월별 활용
        #b. TB는 그냥 가져온다. (월별이 아님)
        tb_월별 = tb[['T1', 'T2', 'T3', 'T4','Company code','통제활동의존', '위험수준', '증감금액', '증감비율', 'Threshold', '분석대상','계정과목코드', '계정과목명', 'CY', 'PY', 'PY1']]

        ### PHW_231019 보완 => PL의 PY는 0으로 변경한다.
        tb_월별['PY'] = np.where(tb['BSPL'] == 'BS', tb['PY'], 0)    

        #c. 당기 GL(피벗) & TB를, 계정과목 KEY로 OUTER JOIN한다.
        avg_bal = pd.merge(gl_월별, tb_월별, on=['계정과목코드', '계정과목명'], how='outer')

        #d. 누적합(cumsum)을 위해 열 순서를 정렬한다.

        # 보완 => 동적으로 구현
        columns = ['T1', 'T2', 'T3', 'T4','Company code','통제활동의존', '위험수준', '증감금액', '증감비율', 'Threshold', '분석대상', '계정과목코드','계정과목명','PY']
        start_column = len(columns) - 1
        
        #TmpList = gl_당기['회계월'].drop_duplicates().to_list() #당기 월수를 추출하고,     :DD231103    
        tc.Set()
        TmpList = gl_당기['회계월'].drop_duplicates()
        if SetGlobal.bDask: TmpList = TmpList.compute()
        tc.Check("gl_당기['회계월'].drop_duplicates()")
        
        TmpList = TmpList.to_list() #list임
        
        TmpList = pd.Series(TmpList).dropna().to_list() #231102 DEBUG, dropna (회계월이 NaN인 경우 삭제)
        TmpList.sort() #오름차순 정렬한다. (누적합을 위해)
        columns = columns + TmpList #합치고,
        columns.append('CY') #마지막에 당기말 포함

        # 누적합을 위한 임시 DF 생성
        avg_bal1 = avg_bal[columns]        
        end_column = len(avg_bal1.columns) -1 #Range는 마지막은 포함하지 않으므로, 그냥 CY까지 선택하면 됨

        # cumsum 수행할 열 선택
        selected_columns = avg_bal1.columns[start_column:end_column]    

        # 선택한 열에 대해서만 cumsum 수행
        avg_bal1.loc[:,selected_columns] = avg_bal1[selected_columns].fillna(0).cumsum(axis=1) #DEBUG

        ################ 저장 ######################

        #a. 월별잔액    
        Filename = f'분석보고서_계정별월별잔액(기초+월별누적증감)_{SetGlobal.ClientNameDate}_{SetGlobal.Level}.xlsx'
        avg_bal1.to_excel(Filename)

        # 안내문 추가
        
        wb = openpyxl.load_workbook(Filename)
        ws = wb.active
        ws.insert_rows(1)
        ws['A1'] = "#계정과목별 각 월말 잔액(기초잔액 + 각 월 누적 증감액) - G/L로부터 추출한 금액"
        ws.column_dimensions.group('B','L',hidden=True)     
        wb.save(Filename)
        wb.close()

        print("월별 말잔 보고서를 생성하였습니다.")

        return tb_월별
        #####################################################################
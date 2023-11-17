########################################################################
# AnalyzeAccounts : 원장 Scan부 => 핵심알고리즘
# v0.0.2 DD 231110
#
# v0.0.2 추상화 리팩토링
########################################################################

import re
import string
import gc

import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import NamedStyle, Font, Alignment, numbers
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import Color, PatternFill, Font, Border
import pandas as pd
import numpy as np
import tqdm
import dask.dataframe as dd
from dask.diagnostics import ProgressBar

from ExcelPar.mod.SetGlobal import SetGlobal
from ExcelPar.mylib.TimeCheck import TimeCheck as tc

class AnalyzeAccounts:

    ########################################################################
    #Define Class var - General
    pbar:tqdm.tqdm
    
    #Define Class var - to decompose
    account_code:str #현재 순환중인 계정과목코드

    df:dd.DataFrame|pd.DataFrame #Raw Full G/L => Use Flag
    df1:dd.DataFrame|pd.DataFrame #Sliced (each account)
    df2:pd.DataFrame #df1 -> Pivot

    wb:Workbook #Created file
    sheet:Worksheet #Created Sheet
    sheet2:Worksheet  #Created Sheet
    sheet3:Worksheet  #Created Sheet

    result_pv:pd.DataFrame #월별 피벗결과
    result1:pd.DataFrame

    추출월:list #대상으로 추출된 월의 List

    ########################################################################
    #ENTRY POINT : 선언부
    ########################################################################
    @classmethod
    def AnalyzeAccounts(cls, gl:dd.DataFrame|pd.DataFrame): #Public

        # List of account codes to analyze
        account_codes = SetGlobal.분석계정과목 #Global
        SetGlobal.검토문장 = []

        cls.account_code = account_codes[0]

        ### 진행률
        cls.pbar = tqdm.tqdm(total=len(account_codes), desc="...")
        cls.pbar.set_description("순환 START")
        ProgressBar().register()

        #i = 0 #FOR DEBUG
        
        for account_code in account_codes: #분석계정과목 전체를 계정별로 순환한다.

            if int(account_code[:8]) <= 35002030: print(account_code); continue
            #i += 1 #FOR DEBUG
            #if i<=18: continue # FOR DEBUG
            #BEGIN
            cls.account_code = account_code

            cls.pbar.set_description(cls.account_code)        
            cls.pbar.update(1)                    

            cls.df = gl #인수로 GL을 받아서 클래스변수에 넣는다.

            cls.df2 = cls.__SetDfSliced()
            cls.__SetCorridor() #범위를 설정한다.

            report_filename = cls.__CreateWorkbook() #Workbook을 생성하고, sheet1을 작성한다.
            cls.__FormatSheet1() #Sheet1을 꾸민다.

            cls.__CreateSheet2() #Sheet2를 만든다.       
            cls.__PivotMonthly() #월별 피벗을 실시하고, 클래스변수로 저장한다.
            cls.__SetSheet2() #피벗 결과를 Sheet2에 찍는다.
            
            cls.__CreateSheet3() #Sheet3을 만든다.
            cls.__SetSheet3() #Sheet3을 입력한다.

            cls.__FormatAllSheets() #일괄 서식을 꾸민다.
            cls.__FormatSheet3()  #시트3 설명을 추가한다.
            
            cls.__SaveWorkbook(report_filename) #Workbook을 저장한다.

            cls.__CreateSentence() #해당 계정 분석문장을 생성하고, 전역변수에 추가한다.

            #CLEAR부
            del cls.df1, cls.df2, cls.result1
            gc.collect()
            
        cls.pbar.close()
        print("계정순환 종료.")

    #####################################################################
    # 이하 정의부
    #####################################################################

    @classmethod
    def __SetDfSliced(cls) -> pd.DataFrame:

        #df1 설정부 : Slicing
        if SetGlobal.Level == 'Detail': #ChangeTBGL에서 이쪽으로 돌아옴
            
            #tc.Set()
            cls.df1 : pd.DataFrame|dd.DataFrame = cls.df[cls.df["Company code"] == cls.account_code] #GL 중 대상 계정과목만 추출 => df1
            #if SetGlobal.bDask: cls.df1 = cls.df1.compute()
            #tc.Check("df1.compute")
            

        elif SetGlobal.Level == 'FSLine':
            #1. 일단 Code 기준으로 추출
            tmp:str = cls.account_code.split("_")[0]

            tc.Set()
            cls.df1 = cls.df[cls.df["FSCode"] == tmp]#.compute #GL 중 대상 계정과목만 추출 => df1
            if SetGlobal.bDask: cls.df1 = cls.df1.compute()
            tc.Check("df1.compute")
        
            #2. 추출 후 가공
            try:
                cls.df1.loc[:,'Company code'] = cls.df1['FSCode'].fillna(0).astype(float).astype(int).astype(str) + "_" + cls.df1['FSName'].astype(str) #DEBUG 231025
            except Exception as e:
                print(e)
                cls.df1.loc[:,'Company code'] = cls.df1['FSCode'].fillna(0).astype(str) + "_" + cls.df1['FSName'].astype(str) #DEBUG 231025
            try:
                cls.df1.loc[:,'계정과목코드'] = cls.df1['FSCode'].fillna(0).astype(float).astype(int).astype(str) #DEBUG 231025
            except Exception as e:
                print(e)
                cls.df1.loc[:,'계정과목코드'] = cls.df1['FSCode'].fillna(0).astype(str) #DEBUG 231025
            cls.df1.loc[:,'계정과목명'] = cls.df1['FSName']            
        else:
            print("전역변수 'Level' 설정 오류")        
        
        #cls.df1 = cls.df[cls.df["Company code"] == cls.account_code] #GL 중 대상 계정과목만 추출 => df1

        d = pd.DataFrame(np.zeros((12, 0))) #[12,0] DF 생성 => d
        d["회계월"] = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]  #d는 1~12 월 Series임
        #df2 = cls.df1.pivot_table(index=["회계월"], columns='연도', values='전표금액', aggfunc='sum').reset_index() #df1을 (행)회계월/(열)연도로 피벗함
        df2_tmp:pd.DataFrame = cls.df[cls.df["Company code"] == cls.account_code].groupby(['회계월','연도'])['전표금액'].sum()        
        tc.Set()
        if SetGlobal.bDask: df2_tmp = df2_tmp.compute()
        tc.Check('Line149')
        df2_tmp = df2_tmp.unstack('연도').reset_index()        
        df2 = pd.merge(d, df2_tmp, on="회계월", how="left").fillna(0) #d & df2 join함. 그냥 df2랑 같은 것임 (1~12월의 12행)        

        #당기 혹은 전기가 없는 에러 수정    
        if 'CY' not in df2.columns:
            df2['CY'] = 0
                # "PY" 열 추가
        if 'PY' not in df2.columns:
            df2['PY'] = 0
            
        df2 = df2[["회계월", "CY", "PY"]]
        return df2

    @classmethod
    def __SetCorridor(cls):
        ########################################################################################
        # 상하단 Logic 변경 # RANGE : NEW LOGIC _ 231023 - 이예린SM
        ########################################################################################

        df_melted = cls.df2.melt(id_vars=['회계월'], value_vars=['CY', 'PY'], var_name='구분', value_name='금액') #Melted Dataframe
        df_mean = df_melted[df_melted['금액'] != 0]["금액"].mean() #Melted df 중 금액이 있는 월의 평균
        df_std = df_melted[df_melted['금액'] != 0]["금액"].std() #표준편차            
        if pd.isna(df_std): df_std = 0 #DEBUG            

        # Z-Score 임계값 설정 (예시: ±2)
        상단기준_Z = 2
        하단기준_Z = -2

        # 상단 기준금액과 하단 기준금액 계산
        상단금액기준 = df_mean + 상단기준_Z * df_std
        하단금액기준 = df_mean + 하단기준_Z * df_std

        cls.df2["상단"] = 상단금액기준
        cls.df2["하단"] = 하단금액기준
        
        cls.df2["당기_분석대상"] = np.where(cls.df2["CY"].abs() <= SetGlobal.De_minimis, "X",
                                np.where((cls.df2["CY"] < cls.df2["상단"]) & (cls.df2["CY"] > cls.df2["하단"]), "X", "O")) ##########수정 10.20 
        cls.df2["전기_분석대상"] = np.where(cls.df2["PY"].abs() <= SetGlobal.De_minimis, "X",
                                np.where((cls.df2["PY"] < cls.df2["상단"]) & (cls.df2["PY"] > cls.df2["하단"]), "X", "O")) ##########수정 10.20 
        
        cls.df2 = cls.df2.set_index('회계월')

    @classmethod
    def __CreateWorkbook(cls) -> str:
        #######################################################################################################
        # Save the analysis report to a separate sheet
        
        # DEBUG. 231019. / to _ (계정과목명에 /가 들어가면 안됨)            
        
        #USE RE
        p = re.compile('[/:<>]')
        account_code_tmp = re.sub(p,"_",cls.account_code)

        report_filename = f"분석보고서_{SetGlobal.ClientNameDate}_{account_code_tmp}_{SetGlobal.Level}.xlsx"
        cls.pbar.set_description(cls.account_code + " 분석보고서 파일을 생성합니다...")
        cls.df2.to_excel(report_filename, sheet_name="분석보고서", startrow=4) #CREATE
        cls.wb = load_workbook(report_filename)
        cls.sheet = cls.wb["분석보고서"]
    
        min_column = cls.wb.active.min_column
        max_column = cls.wb.active.max_column
        min_row = cls.wb.active.min_row
        max_row = cls.wb.active.max_row        

        #######################################################################################################
        # 라인 차트 생성
        line_chart = LineChart()

        data = Reference(cls.sheet, min_col=min_column+1, max_col=max_column-2, min_row=min_row, max_row=max_row) # 111718 선급금'!$B$5:$E$17
        categories = Reference(cls.sheet, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row) # 111718 선급금'!$A$6:$A$17
    
        line_chart.add_data(data, titles_from_data=True)
        line_chart.set_categories(categories)
        cls.sheet.add_chart(line_chart, "B19") #location chart

        line_chart.title = '계정분석'
        line_chart.style = 2 #choose the chart style
        #######################################################################################################    
        return report_filename

    @classmethod
    def __FormatSheet1(cls):
        # #######################################################################################################

        # formatting the report
        cls.sheet['A1'] = '계정분석 보고서'
        cls.sheet['A2'] = cls.account_code
        cls.sheet['A1'].font = Font('Arial', bold=True, size=18)
        cls.sheet['A2'].font = Font('Arial', bold=True, size=12)

        cls.sheet['A3'] = 'CY/PY는 해당월 순증감액입니다. (Not 잔액)'
        cls.sheet['A4'] = '전기~당기 월별순증감액 μ±2σ(=Z-score ±2)가 상하단 정상Range이며, 해당월 순증감액이 Range를 벗어나는 경우 분석대상입니다.'            
    
        #######################################################################################################
    @classmethod
    def __CreateSheet2(cls):
        # Create a new sheet '분석보고서_원장_요약'
        cls.pbar.set_description(cls.account_code + " 분석보고서_원장_요약 시트를 생성합니다...")

        cls.wb.create_sheet("분석보고서_원장_요약", 1)
        cls.sheet2 = cls.wb['분석보고서_원장_요약']

    @classmethod
    def __PivotMonthly(cls):
        df3 = cls.df2.reset_index()
        cls.추출월 = df3[(df3["당기_분석대상"] == "O") | (df3["전기_분석대상"] == "O")]["회계월"].to_list()
        
        if len(cls.추출월) > 0:
        
            # result = pd.DataFrame()
            # for i in cls.추출월:
            #     result = pd.concat([result, cls.df1[cls.df1["회계월"] == i]])  
            # result_pv = pd.pivot_table(result, index = ['회계월','거래처코드'],columns = ['연도'], values = ['전표금액'], aggfunc = 'sum').reset_index().fillna(0)
            
            result = pd.DataFrame()
            for i in cls.추출월: #231113 DEBUG : 아래로 대체
                #result = pd.concat([result, cls.df1[cls.df1["회계월"] == i]])                  
                dfTmp = cls.df[(cls.df["Company code"] == cls.account_code) & (cls.df["회계월"] == i)]
                tc.Set()
                if SetGlobal.bDask: dfTmp = dfTmp.compute()
                tc.Check('cls.df[(cls.df["Company code"] == cls.account_code) & (cls.df["회계월"] == i)]')
                result = pd.concat([result,dfTmp])
            result_pv = pd.pivot_table(result, index = ['회계월','거래처코드'],columns = ['연도'], values = ['전표금액'], aggfunc = 'sum').reset_index().fillna(0)

            # result = pd.DataFrame()
            # for i in cls.추출월:
            #     dfTmp1 = cls.df[cls.df["Company code"] == cls.account_code]
            #     dfTmp2 = dfTmp1[dfTmp1["회계월"] == i]
            #     dfTmp3 = dfTmp2.groupby(["회계월","거래처코드","연도"])["전표금액"].sum()
            #     tc.Set()
            #     dfTmp4 = dfTmp3.compute()
            #     dfTmp5 = dfTmp4.unstack().reset_index()
            #     tc.Check("line 275")                
            #     result = pd.concat([result,dfTmp5])
            # result_pv = result.fillna(0)

            #당기 혹은 전기가 없는 에러 수정
            if ( '전표금액', 'CY') not in result_pv.columns:
                result_pv[( '전표금액', 'CY')] = 0
            # "PY" 열 추가
            if ( '전표금액', 'PY') not in result_pv.columns:
                result_pv[( '전표금액', 'PY')] = 0

            result_pv = result_pv[[(  '회계월',   ''), ('거래처코드',   ''), ( '전표금액', 'CY'), ( '전표금액', 'PY')]]

            result_pv.columns = ["회계월", "거래처코드", "당기", "전기"]
            result_pv["차이금액"] = result_pv["당기"] - result_pv["전기"]
            result_pv["차이금액(절대값)"] = result_pv["차이금액"].abs()
            result_pv["Rank"] = result_pv.groupby(["회계월"])['차이금액(절대값)'].rank(method='min', ascending=False)
            
            cls.result_pv = result_pv #클래스변수로 정의

            result["월합산"] = result.groupby(["회계월","연도"])["전표금액"].transform("sum")
            result["설명율"] = result["전표금액"]/result["월합산"]
            result["기여도"] = result["설명율"].abs()
            result1 = result[result["기여도"] > 0.1].reset_index(drop=True)

            result1["전표금액(절대값)"] = result1["전표금액"].abs()
            result1["Rank"] = result1.groupby(["연도","회계월"])['전표금액(절대값)'].rank(method='min', ascending=False)
            result1 = result1[result1["Rank"] <= 3] 

            cls.result1 = result1 #클래스변수로 정의
            
        else:
            
            data = {
            '회계월': [0],
            '거래처코드': [0],
            '당기': [0],
            '전기': [0],
            '차이금액': [0],
            '차이금액(절대값)': [0],
            'Rank': [0]
        }

            cls.result_pv = pd.DataFrame(data)
            
            data1 = {
            '전표번호': [0],
            '전기일자': [0],
            '연도': [0],
            '회계월': [0],
            '거래처코드': [0],
            '전표적요상세': [0],
            '전표금액': [0],
            '차변금액': [0],
            '대변금액': [0],
            '계정과목코드': [0],
            '계정과목명': [0],
            'Tier 1': [0],
            'Tier 2': [0],
            'Tier 3': [0],
            'Tier 4': [0],
            'Company code': [0],
            '월합산': [0],
            '설명율': [0],
            '기여도': [0],
            '전표금액(절대값)': [0],
            'Rank': [0]
        }
            
            cls.result1 = pd.DataFrame(data1)        

    @classmethod
    def __SetSheet2(cls):
        #######################################################################################################
        # Save the data from 'result_pv' DataFrame to 'sheet2' with headers and border lines
        header_row2 = cls.result_pv.columns.tolist()
        for r_idx, row in enumerate(cls.result_pv.iterrows(), start=5):
            for c_idx, value in enumerate(row[1], start=1):
                cell = cls.sheet2.cell(row=r_idx, column=c_idx, value=value)
        
        for r_idx, header in enumerate(header_row2, start=1):
            cell = cls.sheet2.cell(row=4, column=r_idx, value=header)
            cell.font = Font(bold=True)  # Make the header text bold
            cell.border = openpyxl.styles.Border(
                bottom=openpyxl.styles.Side(style="thin"),
                top=openpyxl.styles.Side(style="thin"),
                left=openpyxl.styles.Side(style="thin"),
                right=openpyxl.styles.Side(style="thin")
            )
                    
        #######################################################################################################
        # formatting the report
        cls.sheet2['A1'] = '분석보고서_원장_요약'
        cls.sheet2['A2'] = '이상변동이 확인된 월을 확인하기 위해 원장변동을 거래처별, 월별로 요약(pivot)합니다.'
        cls.sheet2['A1'].font = Font('Arial', bold=True, size=18)
        cls.sheet2['A2'].font = Font('Arial', bold=True, size=12)
        
        ##############################       
        

        #######################################################################################################

    @classmethod
    def __CreateSheet3(cls):
        cls.pbar.set_description(cls.account_code + " 분석보고서_원장 시트를 생성합니다...")
        # Create a new sheet '분석보고서_원장'
        cls.wb.create_sheet("분석보고서_원장", 2)  # Note: I changed the index to 2 to place it after 'sheet2'
        cls.sheet3 = cls.wb['분석보고서_원장']

    @classmethod
    def __SetSheet3(cls):
        # Save the data from 'result' DataFrame to 'sheet3' with headers and border lines
        header_row3 = cls.result1.columns.tolist()
        for r_idx, row in enumerate(cls.result1.iterrows(), start=5):
            for c_idx, value in enumerate(row[1], start=1):
                cell = cls.sheet3.cell(row=r_idx, column=c_idx, value=value)
        
        for r_idx, header in enumerate(header_row3, start=1):
            cell = cls.sheet3.cell(row=4, column=r_idx, value=header)
            cell.font = Font(bold=True)  # Make the header text bold
            cell.border = openpyxl.styles.Border(
                bottom=openpyxl.styles.Side(style="thin"),
                top=openpyxl.styles.Side(style="thin"),
                left=openpyxl.styles.Side(style="thin"),
                right=openpyxl.styles.Side(style="thin")
            )

    @classmethod
    def __FormatAllSheets(cls):
        #######################################################################################################
        # Define cell format as a DifferentialStyle object
        fill_red = PatternFill(start_color='EE1111', end_color='EE1111', fill_type='solid')
        font = Font(bold=True, color='FFFFFF')
        dxf=DifferentialStyle(font=font ,fill=fill_red)

        # Generation of Rule object
        # Arg:rank sets the rank, arg:bottom selects the upper or lower rank, and arg:percent selects the rank or %.
        rule = Rule(type='top10', rank=10, bottom=False, percent=True, dxf=dxf)
        
        #wb['분석보고서_원장'].conditional_formatting.add("S5:S200", rule)
        cls.wb['분석보고서_원장'].conditional_formatting.add("T5:T200", rule)
        cls.wb['분석보고서_원장_요약'].conditional_formatting.add("F5:F200", rule)
        

        #https://www.shibutan-bloomers.com/python_libraly_openpyxl-7_en/5421/
        
        #max_row = wb['분석보고서_원장'].active.max_row
        for row in range(1, 200):
            cls.wb['분석보고서']["B{}".format(row)].number_format = '#,##0'
            cls.wb['분석보고서']["C{}".format(row)].number_format = '#,##0'
            cls.wb['분석보고서']["D{}".format(row)].number_format = '#,##0'
            cls.wb['분석보고서']["E{}".format(row)].number_format = '#,##0'
            cls.wb['분석보고서']["F{}".format(row)].number_format = '#,##0'
            cls.wb['분석보고서']["G{}".format(row)].number_format = '#,##0'
            
            cls.wb['분석보고서'].column_dimensions["B"].width = 15
            cls.wb['분석보고서'].column_dimensions["C"].width = 15
            cls.wb['분석보고서'].column_dimensions["D"].width = 15
            cls.wb['분석보고서'].column_dimensions["E"].width = 15
            cls.wb['분석보고서'].column_dimensions["F"].width = 15
            cls.wb['분석보고서'].column_dimensions["G"].width = 15
            cls.wb['분석보고서'].column_dimensions["H"].width = 13
            cls.wb['분석보고서'].column_dimensions["I"].width = 13
            
            
            cls.wb['분석보고서_원장_요약']["C{}".format(row)].number_format = '#,##0'
            cls.wb['분석보고서_원장_요약']["D{}".format(row)].number_format = '#,##0'
            cls.wb['분석보고서_원장_요약']["E{}".format(row)].number_format = '#,##0'
            cls.wb['분석보고서_원장_요약']["F{}".format(row)].number_format = '#,##0'
            cls.wb['분석보고서_원장_요약']["G{}".format(row)].number_format = '#,##0'
            
            cls.wb['분석보고서_원장_요약'].column_dimensions["C"].width = 15
            cls.wb['분석보고서_원장_요약'].column_dimensions["D"].width = 15
            cls.wb['분석보고서_원장_요약'].column_dimensions["E"].width = 15
            cls.wb['분석보고서_원장_요약'].column_dimensions["F"].width = 15
            cls.wb['분석보고서_원장_요약'].column_dimensions["G"].width = 15
            
            
            cls.wb['분석보고서_원장']["G{}".format(row)].number_format = '#,##0'
            cls.wb['분석보고서_원장']["H{}".format(row)].number_format = '#,##0'
            cls.wb['분석보고서_원장']["I{}".format(row)].number_format = '#,##0'
            cls.wb['분석보고서_원장']["Q{}".format(row)].number_format = '#,##0'
            cls.wb['분석보고서_원장']["R{}".format(row)].number_format = '#,##0.00'
            cls.wb['분석보고서_원장']["S{}".format(row)].number_format = '#,##0.00'
            cls.wb['분석보고서_원장']["T{}".format(row)].number_format = '#,##0'
            
            cls.wb['분석보고서_원장'].column_dimensions["B"].width = 18
            cls.wb['분석보고서_원장'].column_dimensions["F"].width = 30
            cls.wb['분석보고서_원장'].column_dimensions["G"].width = 15
            cls.wb['분석보고서_원장'].column_dimensions["H"].width = 15
            cls.wb['분석보고서_원장'].column_dimensions["I"].width = 15
            cls.wb['분석보고서_원장'].column_dimensions["Q"].width = 15
            cls.wb['분석보고서_원장'].column_dimensions["T"].width = 15
                
    @classmethod
    def __FormatSheet3(cls):
        # formatting the report
        cls.sheet3['A1'] = '분석보고서_원장'
        cls.sheet3['A2'] = '분석대상 계정의 원장 주요한 변동을 추출합니다(월별증감액합계대비증감액(기여도)이 10%이상이며, 연도별, 월별로 증감금액(절대값)이 상위 3위 이내).'
        cls.sheet3['A3'] = '설명율 = 전표금액/월합산액, 기여도 = abs(설명율)' #231023 추가
        cls.sheet3['A1'].font = Font('Arial', bold=True, size=18)
        cls.sheet3['A2'].font = Font('Arial', bold=True, size=12)
        
        #sheet3.column_dimensions.group('J','S',hidden=True)     # B 부터 D 열 숨기기
        cls.sheet3.column_dimensions.group('J','U',hidden=True)     # Hardcoded        

    @classmethod
    def __SaveWorkbook(cls, report_filename:str):
        cls.pbar.set_description(cls.account_code + " 분석보고서 파일을 저장합니다.")

        #######################################################################################################
        # Save the workbook
        cls.wb.save(report_filename)
                        
        #######################################################################################################

    @classmethod
    def __CreateSentence(cls):
        #분석적검토##################################################################################################################        
        cls.pbar.set_description(cls.account_code + " 분석보고서 문장을 생성합니다...")
        
        # 1.전반적인 개요
        분석보고서_문장 = f"당기 및 전기에 평균 변동금액대비 이상변동을 보인 월은 {', '.join(map(str, cls.추출월))}, 총 {len(cls.추출월)}개 입니다."
        # 2.거래처별, 월별 요약
        월별_최대_거래처 = cls.result_pv.groupby('회계월').apply(lambda x: x.nlargest(1, '차이금액(절대값)'))
        월별_최대_거래처.reset_index(drop=True, inplace=True)
        분석보고서_원장_요약_문장 = "해당월과 거래처로 집계한 내역에서 전기대비 가장 크게 증가한 당기 내용은 아래와 같습니다. \n"
        # 결과 문장 생성
        for index, row in 월별_최대_거래처.iterrows():
            월 = row['회계월']
            거래처 = row['거래처코드']
            금액 = int(row['차이금액'] / 1e6)  # 백만원 단위로 변환
            분석보고서_원장_요약_문장 += f"{월}월, 거래처코드: '{거래처}', {금액}백만원"
            if index < len(월별_최대_거래처) - 1:
                분석보고서_원장_요약_문장 += "와 "
        분석보고서_원장_요약_문장 += "입니다."
        # 3. 원장세부내역 중 주요한 증감
        월별_최대_거래 = cls.result1.groupby(['회계월']).apply(lambda x: x.nlargest(1, '전표금액(절대값)'))
        월별_최대_거래.reset_index(drop=True, inplace=True)
        분석보고서_원장_문장 = "원장을 스캔한 결과 가장 크게 증가(혹은 감소)한 월, 전표번호, 거래처, 금액은 아래와 같습니다.\n"
        for index, row in 월별_최대_거래.iterrows():
            연도 = row["연도"]
            월 = row['회계월']
            전표번호 = row['전표번호']
            거래처 = row['거래처코드']
            금액 = int(row['전표금액'] / 1e6)  # 백만원 단위로 변환
            전표적요 = row['전표적요상세']
            분석보고서_원장_문장 += f"{연도}, {월}월, 전표번호: '{전표번호}', 거래처코드: '{거래처}', {금액}백만원, 세부증가내역은 '{전표적요}'"
            if index < len(월별_최대_거래처) - 1:
                분석보고서_원장_문장 += "와 \n"
        분석보고서_원장_문장 += "입니다."
        # 결과 문장 합치기
        
        if len(cls.추출월) > 0:
        
            결과_문장 = f"{cls.account_code}에서 증가(혹은 감소)한 세부 내역은 아래와 같습니다.\n"
            결과_문장 += '① ' + 분석보고서_문장 + '\n' + '② ' + 분석보고서_원장_요약_문장 + '\n' + '③ ' + 분석보고서_원장_문장
        
        else:
            결과_문장 = f"{cls.account_code}에서 당기 및 전기에 평균 변동금액대비 이상변동을 보인 월은 없으므로 세부 분석은 수행하지 않습니다."        

        SetGlobal.검토문장.append(결과_문장)